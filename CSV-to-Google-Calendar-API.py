import openpyxl as xl
from openpyxl import Workbook, load_workbook
import pyexcel as p
from datetime import datetime
import pandas as pd
import os
from win32com.client import Dispatch
import pickle
import datetime
from collections import namedtuple
from google_auth_oauthlib.flow import Flow, InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from google.auth.transport.requests import Request


# create api service function
def create_service(client_secret_file, api_name, api_version, *scopes, prefix=''):
	CLIENT_SECRET_FILE = client_secret_file
	API_SERVICE_NAME = api_name
	API_VERSION = api_version
	SCOPES = [scope for scope in scopes[0]]
	cred = None
	working_dir = os.getcwd()
	token_dir = 'token files'
	pickle_file = f'token_{API_SERVICE_NAME}_{API_VERSION}{prefix}.pickle'

	# check if token dir exists first, if not, create the folder
	if not os.path.exists(os.path.join(working_dir, token_dir)):
		os.mkdir(os.path.join(working_dir, token_dir))
	if os.path.exists(os.path.join(working_dir, token_dir, pickle_file)):
		with open(os.path.join(working_dir, token_dir, pickle_file), 'rb') as token:
			cred = pickle.load(token)

	if not cred or not cred.valid:
		if cred and cred.expired and cred.refresh_token:
			cred.refresh(Request())
		else:
			flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
			cred = flow.run_local_server()

		with open(os.path.join(working_dir, token_dir, pickle_file), 'wb') as token:
			pickle.dump(cred, token)
	try:
		service = build(API_SERVICE_NAME, API_VERSION, credentials=cred)
		print(API_SERVICE_NAME, API_VERSION, 'service created successfully')
		return service
	except Exception as e:
		print(e)
		print(f'Failed to create service instance for {API_SERVICE_NAME}')
		os.remove(os.path.join(working_dir, token_dir, pickle_file))
		return None


# convert datetime function
def convert_to_RFC_datetime(start_year=1900, start_month=1, start_day=1, hour=0, minute=0):
	dt = datetime.datetime(start_year, start_month, start_day, hour, minute, 0).isoformat() + 'Z'
	return dt


# launch api
CLIENT_SECRET_FILE = "client_secret.json"      ################ do not forget to download OAuth json file and put it in current directory!!!!!!!!!!
API_NAME = 'calendar'
API_VERSION = 'v3'
SCOPES = ['https://www.googleapis.com/auth/calendar']

service = create_service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES)


# list calendars
calendar_list = service.calendarList().list(pageToken=None).execute()


# delete already existing calendar               ### COMPLETELY OPTIONAL - feel free to comment!!!!
for calendar_list_entry in calendar_list['items']:  
    if 'CALENDAR_NAME' in calendar_list_entry['summary']:
        id = calendar_list_entry['id'] 
        service.calendars().delete(calendarId=id).execute()


# create new calendar
calendar_body = {
    'summary': 'CALENDAR_NAME',
    'timeZone': 'America/Sao_Paulo',    ### YOU CAN CHECK TIMEZONES TABLE -> https://en.wikipedia.org/wiki/List_of_tz_database_time_zones
}
service.calendars().insert(body=calendar_body).execute()


# list calendars
calendar_list = service.calendarList().list(pageToken=None).execute()


# get calendar id
for calendar_list_entry in calendar_list['items']:
    if 'CALENDAR_NAME' in calendar_list_entry['summary']:
        id = calendar_list_entry['id'] 



# insert events to google calendar function
def insert_events(color):
    # validate if it is an all-day event or not
    def is_all_day_event():
        if "FALSE" in all_day_event[i]:
            return False
        else:
            return True
    x = 0
    for i in range(0,max_rows):
        if is_all_day_event():
            all_day_event_true_start.append("{}-{}-{}".format(start_year[i],start_month[i],start_day[i]))
            all_day_event_true_end.append("{}-{}-{}".format(end_year[i],end_month[i],end_day[i]))
            event_request_body = {
                'start':{
                    'date': all_day_event_true_start[x],
                    'timeZone': 'America/Sao_Paulo',
                },
                'end':{
                    'date': all_day_event_true_end[x],
                    'timeZone': 'America/Sao_Paulo',
                },
                'summary': subject[i],
                'description': description[i],
                'location': location[i],
                'colorId': color,
                'visibility': is_private[i]
                #'attendees':[
                #    {
                #        'email': '',
                #        'optional': False,
                #        'responseStatus': 'accepted',
                #    }
                #],
                #'reminders': {
                #    'useDefault': False,
                #    'overrides':[
                #        {'method': 'email', 'minutes': 30},
                #    ]
                #}
            }
            service.events().insert(calendarId=id, body=event_request_body).execute()
            x += 1

        else:
            adjust_timezone = 3 # (this is for UTC-3)         ### CHANGE TO YOUR OWN TIMEZONE, COULD BE -n, +n, or none, BASED ON UTC TIME (0)
            event_request_body = {
            'start':{
                'dateTime': convert_to_RFC_datetime(int(start_year[i]), int(start_month[i]), int(start_day[i]), int(fstart_time[i]) + adjust_timezone, 0),
                'timeZone': 'America/Sao_Paulo',
            },
            'end':{
                'dateTime': convert_to_RFC_datetime(int(end_year[i]), int(end_month[i]), int(end_day[i]), int(fend_time[i]) + adjust_timezone, 0),
                'timeZone': 'America/Sao_Paulo',
            },
            'summary': subject[i],
            'description': description[i],
            'location': location[i],
            'colorId': color,
            'visibility': is_private[i]
            #'attendees':[
            #    {
            #        'email': '',
            #        'optional': False,
            #        'responseStatus': 'accepted',
            #    }
            #],
            #'reminders': {
            #    'useDefault': False,
            #    'overrides':[
            #        {'method': 'email', 'minutes': 30},
            #    ]
            #}
            }
            service.events().insert(calendarId=id, body=event_request_body).execute()

all_day_event_true_start = []
all_day_event_true_end = []



# convert csv file to excel
if not os.path.exists('excel_file.xlsx'):
    read_file = pd.read_csv('csv_file.csv')                   # YOU MAY NEED TO DECLARE THE SPECIFIC ENCODING -> https://docs.python.org/3.7/library/codecs.html#standard-encodings
    read_file.to_excel('excel_file.xlsx', index = None, header = True)



# load xlsx file containing the events
wb = load_workbook('excel_file.xlsx')
ws = wb.active


# delete header row
ws.delete_rows(1)

# row number variable
max_rows = ws.max_row


# copy subject
subject = []
for i in range(1,max_rows+1):
    subject.append(ws.cell(row = i, column = 1).value)



## FORMATTING START DATE
# copy start date
start_date = []
for i in range(1,max_rows+1):
    start_date.append(ws.cell(row = i, column = 2).value)

# append lists
start_month = []
start_day = []
start_year = []

# format start month, day & year
for i in range(0,max_rows):
    start_month.append(start_date[i])
    start_day.append(start_date[i])
    start_year.append(start_date[i])

start_month = [x[:-6] for x in start_month]
start_day = [x[3:-3] for x in start_day]
start_year = ['20' + x[6:] for x in start_year]



## FORMATTING START TIME
# copy start time
start_time = []
for i in range(1,max_rows+1):
    start_time.append(ws.cell(row = i, column = 3).value)

# format start time
fstart_time = []
for i in range(0,max_rows):
    fstart_time.append(start_time[i])
fstart_time = [x[:-3] for x in fstart_time]




## FORMATTING END DATE
# copy end date
end_date = []
for i in range(1,max_rows+1):
    end_date.append(ws.cell(row = i, column = 4).value)

# append lists
end_month = []
end_day = []
end_year = []

# format end month, day & year
for i in range(0,max_rows):
    end_month.append(end_date[i])
    end_day.append(end_date[i])
    end_year.append(end_date[i])

end_month = [x[:-6] for x in end_month]
end_day = [x[3:-3] for x in end_day]
end_year = ['20' + x[6:] for x in end_year]




## FORMATTING END TIME
# copy end time
end_time = []
for i in range(1,max_rows+1):
    end_time.append(ws.cell(row = i, column = 5).value)

# format end time
fend_time = []
for i in range(0,max_rows):
    fend_time.append(end_time[i])
fend_time = [x[:-3] for x in fend_time]




# copy all-day event 
all_day_event = []
for i in range(1,max_rows+1):
    all_day_event.append(ws.cell(row = i, column = 6).value)

# padronize all possible letter case options (by uppercasing them all)
for i in range(0,max_rows):
    all_day_event[i] = all_day_event[i].upper()



# copy description
description = []
for i in range(1,max_rows+1):
    description.append(ws.cell(row = i, column = 7).value)



# copy location
location = []
for i in range(1,max_rows+1):
    location.append(ws.cell(row = i, column = 8).value)



# copy private
is_private = []
for i in range(1,max_rows+1):
    is_private.append(ws.cell(row = i, column = 9).value)

# adapt private to fit json
for i in range(0,max_rows):
    is_private[i] = is_private[i].upper()
    if "TRUE" in is_private[i]:
        is_private[i] = "private"
    else:
        is_private[i] = "default"


# insert events(color) ---> check available colors here => https://lukeboyle.com/blog/posts/google-calendar-api-color-id
insert_events(11)    